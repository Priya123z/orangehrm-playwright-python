from pathlib import Path

from typing import Any

from openpyxl.reader.excel import load_workbook

from utils.readers.base_reader import BaseReader


class ExcelReader(BaseReader):


    def validate(self,filepath: Path,**kwargs: Any) -> None:

        if not filepath.exists():
            raise FileNotFoundError(
                f"File '{filepath}' not found.")

        if filepath.suffix.lower() not in [".xlsx", ".xls"]:
            raise ValueError(
                f"Given file: '{filepath}' is not an excel file.")

        if "sheet_name" not in kwargs:
            raise ValueError(
                "sheet_name is required when reading Excel files."
            )


    def read(self,filepath:Path, **kwargs: Any) -> list[dict]:
        """
        Reads Excel workbooks and returns data
        in a standardized List[Dict] format.
        """

        self.validate(filepath, **kwargs)
        workbook = load_workbook(filepath)
        sheet_name = kwargs["sheet_name"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found.")
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Worksheet {sheet_name} is empty")
        headers = rows[0]
        data = []
        for row in rows[1:]:
            data.append(dict(zip(headers, row)))
        try:
            return data
        finally:
            workbook.close()
